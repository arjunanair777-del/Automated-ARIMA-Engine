"""
Top 10 Diseases in the World – Excel Dashboard Generator
=========================================================
Generates an Excel workbook (top10_diseases_dashboard.xlsx) containing:
  • **Data** sheet   – disease statistics with attractive table formatting
  • **Dashboard** sheet – bar chart, pie chart, and summary KPI cards

Data is based on WHO Global Health Estimates (approximate annual figures).

Usage
-----
    python top10_diseases_dashboard.py          # writes to current directory
    python top10_diseases_dashboard.py out.xlsx  # custom output path
"""

import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Disease data (approximate annual figures, WHO Global Health Estimates)
# ---------------------------------------------------------------------------
DISEASES = [
    {
        "rank": 1,
        "disease": "Ischaemic Heart Disease",
        "category": "Cardiovascular",
        "annual_deaths": 8_900_000,
        "prevalence_millions": 244,
        "trend": "Stable",
        "primary_region": "Global",
    },
    {
        "rank": 2,
        "disease": "Stroke",
        "category": "Cardiovascular",
        "annual_deaths": 6_600_000,
        "prevalence_millions": 101,
        "trend": "Increasing",
        "primary_region": "East Asia",
    },
    {
        "rank": 3,
        "disease": "COPD",
        "category": "Respiratory",
        "annual_deaths": 3_500_000,
        "prevalence_millions": 480,
        "trend": "Increasing",
        "primary_region": "South-East Asia",
    },
    {
        "rank": 4,
        "disease": "Lower Respiratory Infections",
        "category": "Infectious",
        "annual_deaths": 2_600_000,
        "prevalence_millions": 489,
        "trend": "Decreasing",
        "primary_region": "Sub-Saharan Africa",
    },
    {
        "rank": 5,
        "disease": "Trachea / Bronchus / Lung Cancer",
        "category": "Cancer",
        "annual_deaths": 1_800_000,
        "prevalence_millions": 2.2,
        "trend": "Increasing",
        "primary_region": "East Asia",
    },
    {
        "rank": 6,
        "disease": "Diabetes Mellitus",
        "category": "Metabolic",
        "annual_deaths": 1_600_000,
        "prevalence_millions": 537,
        "trend": "Increasing",
        "primary_region": "Global",
    },
    {
        "rank": 7,
        "disease": "Alzheimer's & Dementias",
        "category": "Neurological",
        "annual_deaths": 1_500_000,
        "prevalence_millions": 55,
        "trend": "Increasing",
        "primary_region": "Europe",
    },
    {
        "rank": 8,
        "disease": "Diarrhoeal Diseases",
        "category": "Infectious",
        "annual_deaths": 1_500_000,
        "prevalence_millions": 1_700,
        "trend": "Decreasing",
        "primary_region": "Sub-Saharan Africa",
    },
    {
        "rank": 9,
        "disease": "Tuberculosis",
        "category": "Infectious",
        "annual_deaths": 1_300_000,
        "prevalence_millions": 10.6,
        "trend": "Decreasing",
        "primary_region": "South-East Asia",
    },
    {
        "rank": 10,
        "disease": "Kidney Diseases",
        "category": "Renal",
        "annual_deaths": 1_300_000,
        "prevalence_millions": 850,
        "trend": "Increasing",
        "primary_region": "Global",
    },
]

HEADERS = [
    "Rank",
    "Disease",
    "Category",
    "Annual Deaths",
    "Prevalence (millions)",
    "Trend",
    "Primary Region",
]

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
DATA_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=20, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2E75B6")
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=18, color="1F4E79")
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="4472C4")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

CATEGORY_COLORS = {
    "Cardiovascular": "D6E4F0",
    "Respiratory": "E2EFDA",
    "Infectious": "FCE4D6",
    "Cancer": "F2DCDB",
    "Metabolic": "FFF2CC",
    "Neurological": "E4DFEC",
    "Renal": "D9E2F3",
}

PIE_COLORS = [
    "2E75B6", "548235", "BF8F00", "C00000", "7030A0",
    "00B0F0", "FFC000", "FF6600", "92D050", "4472C4",
]


def _auto_fit_column_widths(ws):
    """Auto-fit rough column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value))
            except TypeError:
                cell_len = 0
            max_len = max(max_len, cell_len)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)


# ---------------------------------------------------------------------------
# Build the Data sheet
# ---------------------------------------------------------------------------
def _build_data_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Data"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Top 10 Diseases in the World"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle row
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = "Annual mortality & prevalence – WHO Global Health Estimates"
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = CENTER
    ws.row_dimensions[2].height = 25

    # Header row (row 4)
    header_row = 4
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 30

    # Data rows
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for i, d in enumerate(DISEASES):
        row = header_row + 1 + i
        values = [
            d["rank"],
            d["disease"],
            d["category"],
            d["annual_deaths"],
            d["prevalence_millions"],
            d["trend"],
            d["primary_region"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx in (1, 3, 5, 6, 7):
                cell.alignment = CENTER
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            else:
                cell.alignment = LEFT

            if i % 2 == 1:
                cell.fill = even_fill

    # Total row
    total_row = header_row + 1 + len(DISEASES)
    ws.merge_cells(f"A{total_row}:C{total_row}")
    total_label = ws.cell(row=total_row, column=1, value="TOTAL ANNUAL DEATHS")
    total_label.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_label.border = THIN_BORDER
    for c in range(2, 4):
        ws.cell(row=total_row, column=c).border = THIN_BORDER

    total_val = ws.cell(
        row=total_row,
        column=4,
        value=sum(d["annual_deaths"] for d in DISEASES),
    )
    total_val.font = Font(name="Calibri", bold=True, size=12, color="C00000")
    total_val.number_format = "#,##0"
    total_val.alignment = Alignment(horizontal="right", vertical="center")
    total_val.border = THIN_BORDER
    for c in range(5, 8):
        ws.cell(row=total_row, column=c).border = THIN_BORDER

    ws.row_dimensions[total_row].height = 25

    # Source note
    note_row = total_row + 2
    ws.merge_cells(f"A{note_row}:G{note_row}")
    note = ws.cell(row=note_row, column=1)
    note.value = "Source: WHO Global Health Estimates (approximate figures)"
    note.font = Font(name="Calibri", italic=True, size=9, color="808080")
    note.alignment = LEFT

    _auto_fit_column_widths(ws)
    # Override a couple of widths for readability
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["G"].width = 22

    # Freeze panes below headers
    ws.freeze_panes = "A5"

    return ws


# ---------------------------------------------------------------------------
# Build the Dashboard sheet
# ---------------------------------------------------------------------------
def _build_dashboard(wb: Workbook):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "2E75B6"

    # ---------- Title ----------
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = "🏥  Global Disease Dashboard – Top 10 Causes of Death"
    t.font = Font(name="Calibri", bold=True, size=22, color="1F4E79")
    t.alignment = CENTER
    t.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:N2")
    st = ws["A2"]
    st.value = "Based on WHO Global Health Estimates (approximate annual figures)"
    st.font = Font(name="Calibri", italic=True, size=11, color="4472C4")
    st.alignment = CENTER

    # ---------- KPI cards (row 4-6) ----------
    total_deaths = sum(d["annual_deaths"] for d in DISEASES)
    avg_prevalence = sum(d["prevalence_millions"] for d in DISEASES) / len(DISEASES)
    increasing = sum(1 for d in DISEASES if d["trend"] == "Increasing")
    categories = len({d["category"] for d in DISEASES})

    kpis = [
        ("Total Deaths / Year", f"{total_deaths:,}", "B"),
        ("Avg Prevalence (M)", f"{avg_prevalence:,.1f}", "E"),
        ("Trends Increasing", str(increasing), "H"),
        ("Disease Categories", str(categories), "K"),
    ]

    kpi_fill = PatternFill(start_color="EAF0F9", end_color="EAF0F9", fill_type="solid")
    kpi_border = Border(
        left=Side(style="medium", color="4472C4"),
        right=Side(style="medium", color="4472C4"),
        top=Side(style="medium", color="4472C4"),
        bottom=Side(style="medium", color="4472C4"),
    )
    for label_text, value_text, start_col in kpis:
        col_idx = ord(start_col) - ord("A") + 1
        end_col = get_column_letter(col_idx + 1)
        # Merge two columns for each KPI card
        ws.merge_cells(f"{start_col}4:{end_col}4")
        ws.merge_cells(f"{start_col}5:{end_col}5")
        for r in (4, 5):
            for c in (col_idx, col_idx + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = kpi_fill
                cell.border = kpi_border
        val_cell = ws.cell(row=4, column=col_idx, value=value_text)
        val_cell.font = KPI_VALUE_FONT
        val_cell.alignment = CENTER
        lbl_cell = ws.cell(row=5, column=col_idx, value=label_text)
        lbl_cell.font = KPI_LABEL_FONT
        lbl_cell.alignment = CENTER

    ws.row_dimensions[4].height = 35
    ws.row_dimensions[5].height = 22

    # ---------- Hidden data block for charts (rows 20-31) ----------
    data_start = 20
    ws.cell(row=data_start, column=1, value="Disease")
    ws.cell(row=data_start, column=2, value="Annual Deaths")
    ws.cell(row=data_start, column=3, value="Prevalence (M)")
    ws.cell(row=data_start, column=4, value="Category")
    for i, d in enumerate(DISEASES):
        r = data_start + 1 + i
        ws.cell(row=r, column=1, value=d["disease"])
        ws.cell(row=r, column=2, value=d["annual_deaths"])
        ws.cell(row=r, column=3, value=d["prevalence_millions"])
        ws.cell(row=r, column=4, value=d["category"])

    data_end = data_start + len(DISEASES)

    # ---------- Bar chart: Annual Deaths ----------
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Annual Deaths by Disease"
    bar.y_axis.title = "Deaths"
    bar.x_axis.title = "Disease"
    bar.y_axis.numFmt = "#,##0"
    bar.width = 28
    bar.height = 16

    data_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    cats_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.shape = 4

    series = bar.series[0]
    series.graphicalProperties.solidFill = "2E75B6"
    series.graphicalProperties.line.solidFill = "1F4E79"

    # Color each bar individually
    for idx, color in enumerate(PIE_COLORS):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)

    bar.legend = None
    ws.add_chart(bar, "A7")

    # ---------- Pie chart: Deaths distribution ----------
    pie = PieChart()
    pie.title = "Share of Deaths by Disease"
    pie.style = 10
    pie.width = 20
    pie.height = 16

    pie_data = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    pie_cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_end)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)

    pie_series = pie.series[0]
    for idx, color in enumerate(PIE_COLORS):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie_series.data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False

    ws.add_chart(pie, "H7")

    # ---------- Category summary table (below charts) ----------
    cat_row = 25
    ws.merge_cells(f"A{cat_row}:F{cat_row}")
    ct = ws.cell(row=cat_row, column=1, value="Deaths by Category")
    ct.font = SUBTITLE_FONT
    ct.alignment = LEFT

    cat_header_row = cat_row + 1
    cat_headers = ["Category", "Diseases", "Total Deaths", "% of Total"]
    for ci, h in enumerate(cat_headers, start=1):
        cell = ws.cell(row=cat_header_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Aggregate by category
    cat_agg = OrderedDict()
    for d in DISEASES:
        cat = d["category"]
        if cat not in cat_agg:
            cat_agg[cat] = {"count": 0, "deaths": 0}
        cat_agg[cat]["count"] += 1
        cat_agg[cat]["deaths"] += d["annual_deaths"]

    even_fill = PatternFill(start_color="EAF0F9", end_color="EAF0F9", fill_type="solid")
    for i, (cat, info) in enumerate(
        sorted(cat_agg.items(), key=lambda x: x[1]["deaths"], reverse=True)
    ):
        r = cat_header_row + 1 + i
        vals = [cat, info["count"], info["deaths"], info["deaths"] / total_deaths]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if ci == 3:
                cell.number_format = "#,##0"
            elif ci == 4:
                cell.number_format = "0.0%"
            if i % 2 == 1:
                cell.fill = even_fill

    # Set column widths
    for col_letter in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["A"].width = 34

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate_dashboard(output_path: str = "top10_diseases_dashboard.xlsx"):
    """Create the Excel workbook and save to *output_path*."""
    wb = Workbook()
    _build_data_sheet(wb)
    _build_dashboard(wb)
    wb.save(output_path)
    print(f"✅  Dashboard saved to: {Path(output_path).resolve()}")
    return output_path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "top10_diseases_dashboard.xlsx"
    generate_dashboard(out)
